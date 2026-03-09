from pathlib import Path
from typing import Any
import logging

import pandas as pd
from openpyxl import Workbook, load_workbook


LOGGER = logging.getLogger(__name__)


def exportar_excel(rows: list[tuple[Any, ...]], cols: list[str], ruta: Path) -> None:
    # Normaliza ruta y extension de salida.
    ruta = Path(ruta)
    if ruta.suffix.lower() == ".xls":
        # Pandas moderno no soporta escritura .xls con engine legacy.
        ruta = ruta.with_suffix(".xlsx")
        LOGGER.warning("Salida .xls no soportada por pandas actual; se exporta como: %s", ruta)

    # Crea carpeta destino si no existe.
    ruta.parent.mkdir(parents=True, exist_ok=True)
    # Convierte datos tabulares a DataFrame.
    df = pd.DataFrame(rows, columns=cols)
    # Exporta a Excel usando openpyxl.
    with pd.ExcelWriter(ruta, engine="openpyxl") as writer:
        df.to_excel(
            writer,
            index=False,
            sheet_name="Reporte",
            freeze_panes=(1, 0),
        )


def exportar_pestana_excel(
    rows: list[tuple[Any, ...]],
    cols: list[str],
    ruta: Path,
    sheet_name: str,
) -> None:
    # Agrega o reemplaza una pestaña en un Excel existente.
    ruta = Path(ruta)
    if ruta.suffix.lower() == ".xls":
        ruta = ruta.with_suffix(".xlsx")
        LOGGER.warning("Salida .xls no soportada por pandas actual; se exporta como: %s", ruta)

    ruta.parent.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(rows, columns=cols)

    if ruta.exists():
        writer = pd.ExcelWriter(
            ruta,
            engine="openpyxl",
            mode="a",
            if_sheet_exists="replace",
        )
    else:
        writer = pd.ExcelWriter(
            ruta,
            engine="openpyxl",
            mode="w",
        )

    with writer:
        df.to_excel(
            writer,
            index=False,
            sheet_name=sheet_name,
            freeze_panes=(1, 0),
        )


def exportar_comparacion(
    resumen: dict[str, int],
    faltantes: list[dict[str, str]],
    diferencias: list[dict[str, str | float]],
    ruta: Path,
    sheet_name: str = "Comparacion",
) -> None:
    # Una sola pestaña con 3 bloques: Resumen | Faltantes | Diferencias.
    ruta = Path(ruta)
    ruta.parent.mkdir(parents=True, exist_ok=True)

    resumen_rows = [
        {"metrica": "sap", "cantidad": resumen.get("sap", 0)},
        {"metrica": "tutati", "cantidad": resumen.get("tutati", 0)},
        {"metrica": "faltan_en_sap", "cantidad": resumen.get("faltan_en_sap", 0)},
        {"metrica": "faltan_en_tutati", "cantidad": resumen.get("faltan_en_tutati", 0)},
    ]
    faltantes_rows = faltantes if faltantes else [{"tipo_faltante": "SIN_FALTANTES", "sap": "", "tutati": "", "fecha": ""}]
    diferencias_rows = diferencias if diferencias else [{
        "u_bot_docentry": "",
        "uid_orders": "",
        "fecha": "",
        "suma_sap": 0.0,
        "total_tutati": 0.0,
        "diferencia": 0.0,
    }]

    if ruta.exists():
        wb = load_workbook(ruta)
    else:
        wb = Workbook()
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    def write_block(
        title: str,
        rows: list[dict[str, Any]],
        start_col: int,
    ) -> int:
        headers = list(rows[0].keys()) if rows else ["sin_datos"]
        ws.cell(row=1, column=start_col, value=title)
        for i, header in enumerate(headers):
            ws.cell(row=2, column=start_col + i, value=header)
        for r, item in enumerate(rows, start=3):
            for c, header in enumerate(headers):
                ws.cell(row=r, column=start_col + c, value=item.get(header, ""))
        return len(headers)

    width_resumen = write_block("RESUMEN", resumen_rows, 1)
    start_faltantes = 1 + width_resumen + 1
    width_faltantes = write_block("FALTANTES", faltantes_rows, start_faltantes)
    start_diferencias = start_faltantes + width_faltantes + 1
    write_block("DIFERENCIAS", diferencias_rows, start_diferencias)

    ws.freeze_panes = "A3"
    wb.save(ruta)
